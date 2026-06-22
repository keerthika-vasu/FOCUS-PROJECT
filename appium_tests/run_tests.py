import os
import sys
import argparse
import base64
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

REPORTS_DIR = os.path.join(os.path.dirname(__file__), "reports")
SCREENSHOTS_DIR = os.path.join(REPORTS_DIR, "screenshots")

os.makedirs(SCREENSHOTS_DIR, exist_ok=True)

class PytestCollector:
    def __init__(self):
        self.results = []

    def pytest_runtest_logreport(self, report):
        if report.when == 'call':
            test_name = report.nodeid.split("::")[-1]
            status = "PASSED" if report.passed else ("FAILED" if report.failed else "SKIPPED")
            screenshot_path = ""
            
            if status == "FAILED":
                potential_screenshot = os.path.join(SCREENSHOTS_DIR, f"{test_name}_failed.png")
                # conftest.py saves screenshots here on failure
                if os.path.exists(potential_screenshot):
                    screenshot_path = os.path.relpath(potential_screenshot, REPORTS_DIR)

            err_msg = ""
            if report.failed:
                if report.longrepr:
                    err_msg = str(report.longreprtext) if hasattr(report, "longreprtext") else str(report.longrepr)
                    if len(err_msg) > 1000:
                        err_msg = err_msg[:1000] + "\n... [TRUNCATED] ..."
            
            self.results.append({
                "Test Name": test_name,
                "Status": status,
                "Duration (s)": round(report.duration, 2),
                "Error Message": err_msg,
                "Screenshot Path": screenshot_path
            })

def generate_excel_report(results, output_path):
    df = pd.DataFrame(results)
    
    total = len(df)
    passed = len(df[df["Status"] == "PASSED"])
    failed = len(df[df["Status"] == "FAILED"])
    skipped = len(df[df["Status"] == "SKIPPED"])
    pass_rate = f"{(passed / total * 100):.1f}%" if total > 0 else "0%"
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "E2E Test Results"
    
    ws.views.sheetView[0].showGridLines = True
    
    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=16, bold=True, color="1B365D")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    bold_font = Font(name=font_family, size=11, bold=True)
    regular_font = Font(name=font_family, size=11)
    
    fill_header = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    fill_pass = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    fill_fail = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    fill_skip = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    
    # Title
    ws.cell(row=1, column=1, value="Appium E2E Test Execution Report").font = title_font
    ws.row_dimensions[1].height = 30
    
    # Summary Info
    summary_data = [
        ("Total Tests Run", total),
        ("Passed Tests", passed),
        ("Failed Tests", failed),
        ("Skipped Tests", skipped),
        ("Pass Rate", pass_rate)
    ]
    
    ws.cell(row=3, column=1, value="Metric").font = header_font
    ws.cell(row=3, column=1).fill = fill_header
    ws.cell(row=3, column=2, value="Value").font = header_font
    ws.cell(row=3, column=2).fill = fill_header
    
    for idx, (metric, val) in enumerate(summary_data, start=4):
        ws.cell(row=idx, column=1, value=metric).font = bold_font
        ws.cell(row=idx, column=2, value=val).font = regular_font
        if metric == "Pass Rate":
            ws.cell(row=idx, column=2).font = Font(name=font_family, size=11, bold=True, color="28A745" if failed == 0 else "DC3545")
            
    thin_border = openpyxl.styles.Border(
        left=openpyxl.styles.Side(style='thin', color='D0D0D0'),
        right=openpyxl.styles.Side(style='thin', color='D0D0D0'),
        top=openpyxl.styles.Side(style='thin', color='D0D0D0'),
        bottom=openpyxl.styles.Side(style='thin', color='D0D0D0')
    )
    for r in range(3, 9):
        for c in range(1, 3):
            ws.cell(row=r, column=c).border = thin_border
            
    # Details Table
    details_start_row = 10
    headers = list(df.columns)
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=details_start_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center")
        
    ws.row_dimensions[details_start_row].height = 25
    
    for row_idx, row_data in enumerate(df.values, start=details_start_row + 1):
        ws.row_dimensions[row_idx].height = 20
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = regular_font
            cell.border = thin_border
            
            col_name = headers[col_idx - 1]
            if col_name == "Status":
                cell.alignment = Alignment(horizontal="center")
                if value == "PASSED":
                    cell.fill = fill_pass
                elif value == "FAILED":
                    cell.fill = fill_fail
                else:
                    cell.fill = fill_skip
            elif col_name == "Duration (s)":
                cell.alignment = Alignment(horizontal="right")
            elif col_name in ["Error Message", "Screenshot Path"]:
                cell.alignment = Alignment(wrap_text=True)
                
    # Auto-fit widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row > 1 and cell.value:
                val_str = str(cell.value)
                if len(val_str) < 50:
                    max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    if "Error Message" in headers:
        err_col_letter = get_column_letter(headers.index("Error Message") + 1)
        ws.column_dimensions[err_col_letter].width = 40
        
    wb.save(output_path)
    print(f"Excel report successfully generated at: {os.path.abspath(output_path)}")

def write_dummy_screenshot(filename):
    # Base64 string of a tiny transparent 1x1 png image
    tiny_png = b'iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mNk+M9QDwADhgGAWjR9awAAAABJRU5ErkJggg=='
    img_data = base64.b64decode(tiny_png)
    screenshot_path = os.path.join(SCREENSHOTS_DIR, filename)
    with open(screenshot_path, 'wb') as f:
        f.write(img_data)
    return os.path.relpath(screenshot_path, REPORTS_DIR)

def write_dummy_log():
    log_path = os.path.join(REPORTS_DIR, "appium.log")
    with open(log_path, 'w', encoding='utf-8') as f:
        f.write("2026-06-19 09:30:00 [INFO] Initializing Appium driver...\n")
        f.write("2026-06-19 09:30:02 [INFO] Appium driver session created successfully.\n")
        f.write("2026-06-19 09:30:05 [INFO] Running test: test_login_fields_visibility...\n")
        f.write("2026-06-19 09:30:10 [INFO] Test test_login_fields_visibility PASSED.\n")
        f.write("2026-06-19 09:30:10 [INFO] Running test: test_login_empty_credentials_error...\n")
        f.write("2026-06-19 09:30:15 [WARNING] Element 'Please enter email' not found immediately, retrying...\n")
        f.write("2026-06-19 09:30:20 [ERROR] Test test_login_empty_credentials_error FAILED.\n")
        f.write("2026-06-19 09:30:20 [ERROR] Screenshot saved to reports/screenshots/test_login_empty_credentials_error_failed.png\n")
        f.write("2026-06-19 09:30:20 [INFO] Running test: test_navigate_to_signup...\n")
        f.write("2026-06-19 09:30:28 [INFO] Test test_navigate_to_signup PASSED.\n")
        f.write("2026-06-19 09:30:30 [INFO] Tearing down Appium driver session...\n")

def run_dry_run():
    print("Executing tests in DRY-RUN mode...")
    write_dummy_log()
    screenshot_rel = write_dummy_screenshot("test_login_empty_credentials_error_failed.png")
    
    mock_results = [
        {
            "Test Name": "test_login_fields_visibility",
            "Status": "PASSED",
            "Duration (s)": 8.24,
            "Error Message": "",
            "Screenshot Path": ""
        },
        {
            "Test Name": "test_login_empty_credentials_error",
            "Status": "FAILED",
            "Duration (s)": 10.45,
            "Error Message": "AssertionError: Expected error message 'Please enter email and password.' was not displayed.\nStack Trace:\n  File 'test_login.py', line 26, in test_login_empty_credentials_error\n    assert error_el is not None",
            "Screenshot Path": screenshot_rel
        },
        {
            "Test Name": "test_navigate_to_signup",
            "Status": "PASSED",
            "Duration (s)": 5.12,
            "Error Message": "",
            "Screenshot Path": ""
        }
    ]
    
    excel_path = os.path.join(REPORTS_DIR, "test_report.xlsx")
    generate_excel_report(mock_results, excel_path)
    print("Dry-run complete. Mock results and reports created successfully.")

def main():
    parser = argparse.ArgumentParser(description="Appium Test Runner & Excel Reporter")
    parser.add_argument("--dry-run", action="store_true", help="Simulate test run and generate reports")
    args = parser.parse_args()

    if args.dry_run:
        run_dry_run()
    else:
        import pytest
        print("Starting live E2E Appium tests using Pytest...")
        collector = PytestCollector()
        test_dir = os.path.dirname(__file__)
        junit_xml = os.path.join(REPORTS_DIR, "junit.xml")
        
        # Discover all test_*.py files in this directory
        print(f"Discovering tests in: {test_dir}")
        
        # Run pytest across all test files with JUnit XML output for CI
        pytest.main([
            "-v",
            test_dir,
            f"--junit-xml={junit_xml}",
            "--tb=short",
        ], plugins=[collector])
        
        if not collector.results:
            print("Error: No test results collected. Verify Appium server is running and configuration is correct.")
            sys.exit(1)
            
        excel_path = os.path.join(REPORTS_DIR, "test_report.xlsx")
        generate_excel_report(collector.results, excel_path)

if __name__ == "__main__":
    main()
